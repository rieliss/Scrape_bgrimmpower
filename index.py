import pandas as pd
import openpyxl as pxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Name of the workbook we'll be using
filename = 'test_wb.xlsx'

# For demo purposes, create a new 1-sheet workbook
# DataFrame to be inserted in the first worksheet
firstMockDF = pd.DataFrame({
    'a': [0,0],
    'b': [0,0]
})
# This creates a new workbook, which will contain only one sheet: sheetA
firstMockDF.to_excel(filename, 'sheetA', index=False)

#######################################
# Load the workbook (i.e. the existing workbook in your case)
excel_book = pxl.load_workbook(filename)

# Create a new worksheet
excel_book.create_sheet('sheetB')
# New data to be written into the new worksheet
secondMockDF = pd.DataFrame({
    'c': [0,0],
    'd': [0,0]
})

# https://stackoverflow.com/a/66534734
rows = dataframe_to_rows(secondMockDF, index=False)
# Work with the new worksheet
ws = excel_book['sheetB']
# For each row
for r_idx, row in enumerate(rows, 1):
    # Write each cell for each column
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Overwrite the workbook, now with two worksheets populated
excel_book.save(filename)